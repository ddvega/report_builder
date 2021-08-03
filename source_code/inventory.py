from collections import defaultdict
import json
import openpyxl
from source_code.functions import *


def create_json(obj, file, archive):
    with open(file, "w") as outfile:
        json.dump(obj, outfile)


def load_json(file):
    try:
        json_obj = open(file, )
        new_dict = json.load(json_obj)
        json_obj.close()
        return new_dict

    except Exception:
        print(f"{file} does not exist.")
        return {}


def update_json_file(file, obj):
    stored = load_json(file)
    merged = stored | obj
    # create_json(
    #     merged, file, f'json/archived/{file}_{datetime.datetime.now()}')
    create_json(merged, file, 'source_data/json/archived')


def update_json_excel(file, file_json, output_box, progress_bar):
    output_box.clear()
    output_msg('Process Initiated.', 1, output_box, progress_bar)
    output_msg(f'Loading {file} file.', 3, output_box, progress_bar)
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    rows = ws.max_row
    obj = {}

    output_msg(f'Reading {file}.', 10, output_box, progress_bar)
    for x in range(1, rows + 1):
        try:
            sku = int(ws['A%s' % x].value)
            val = str(ws['B%s' % x].value)
            obj[sku] = val
            output_box.append(f'SKU({sku}) has been updated to {val}')
        except Exception:
            output_msg(f'{file} is empty. No DATA to extract.', 100, output_box, progress_bar)
            return ""

    output_msg('Converting Dictionary to JSON file.', 60, output_box, progress_bar)
    new_json = json.dumps(obj)

    output_msg(f'Updating {file_json}.', 80, output_box, progress_bar)
    update_json_file(file_json, json.loads(new_json))

    output_msg('Process Complete.', 100, output_box, progress_bar)


class Inventory:
    def __init__(self):
        self.cols = ['A', 'B', 'C', 'D', 'E',
                     'F', 'G', 'H', 'I', 'K', 'L', 'M']
        self.newRows = []
        self.data_from_order_guide = defaultdict(list)
        self.section_names = {}
        self.price_json = None
        self.pack_size_json = None

    def import_from_order_guide(self, output_box, progress_bar):
        output_box.clear()
        output_msg('Process Initiated.', 1, output_box, progress_bar)
        file = "source_data/xlsx/orderguide.xlsx"
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = ws.max_row

        output_msg('Extracting Prices.', 5, output_box, progress_bar)
        for x in range(1, rows):
            for col in self.cols:
                val = getPrices(x, ws, col)
                if val:
                    self.data_from_order_guide[x].append(val)
                    self.newRows.append(x)
                    break

        output_msg('Extracting Sizes.', 20, output_box, progress_bar)
        for x in self.newRows:
            for col in self.cols:
                val = getSize(x, ws, col)
                if val:
                    self.data_from_order_guide[x].append(val)
                    break

        output_msg('Extracting SKUs.', 40, output_box, progress_bar)
        for x in self.newRows:
            for col in self.cols:
                val = getSkus(x, ws, col)
                if val:
                    self.data_from_order_guide[x].append(val)
                    break

        output_msg('Building Dictionaries.', 60, output_box, progress_bar)
        og_prices = {}
        og_pack_sizes = {}
        for key, val in self.data_from_order_guide.items():
            if len(val) == 3:
                og_prices[val[2]] = val[0]
                output_box.append(f'Price Updated for sku({val[2]}) to ${val[0]}')
                og_pack_sizes[val[2]] = val[1]
                output_box.append(f'Pack Size Updated for sku({val[2]}) to {val[0]}')

        output_msg('Converting Dictionaries to JSON.', 70, output_box, progress_bar)
        self.price_json = json.dumps(og_prices)
        self.pack_size_json = json.dumps(og_pack_sizes)

        output_msg('Updating existing JSON files.', 80, output_box, progress_bar)
        update_json_file('source_data/json/prices.json', json.loads(self.price_json))
        update_json_file('source_data/json/pack_sizes.json',
                         json.loads(self.pack_size_json))
        output_msg('Process Complete.', 100, output_box, progress_bar)

    def excel_to_sect_names_json(self, output_box, progress_bar):
        update_json_excel('source_data/xlsx/sect_names.xlsx', 'source_data/json/sect_names.json', output_box,
                          progress_bar)

    def excel_to_pack_sizes_json(self, output_box, progress_bar):
        update_json_excel('source_data/xlsx/pack_sizes.xlsx', 'source_data/json/pack_sizes.json', output_box,
                          progress_bar)

    def excel_to_prices_json(self, output_box, progress_bar):
        update_json_excel('source_data/xlsx/prices.xlsx', 'source_data/json/prices.json', output_box, progress_bar)

    def load_prices(self):
        return load_json('source_data/json/prices.json')

    def load_pack_sizes(self):
        return load_json('source_data/json/pack_sizes.json')

    def load_sec_names(self):
        return load_json('source_data/json/sect_names.json')

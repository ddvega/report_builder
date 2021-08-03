import openpyxl
from collections import defaultdict
import json
import pandas as pd
import math
from source_code.functions import *


class Sales:
    def __init__(self, price_list=None, pack_sizes=None, sect_names=None, output_box=None, progress_bar=None):
        self.data = defaultdict(list)
        self.output_box = output_box
        self.progress_bar = progress_bar
        self.price_list = price_list
        self.pack_sizes = pack_sizes
        self.sect_names = sect_names
        self.report_start_date = ''
        self.report_end_date = ''
        self.TOS = []

    def clean_sales(self):
        output_msg("Data cleaning started.", 6, self.output_box, self.progress_bar)
        file = "source_data/xlsx/hsog.xlsx"
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = ws.max_row

        self.report_start_date = get_report_date(ws, 'A%s')
        self.report_end_date = self.report_start_date + \
                               datetime.timedelta(days=13)

        move_columns(ws, rows)
        filter_for_skus(ws, rows)

        output_msg("Building dictionary from cleaned data.", 10, self.output_box, self.progress_bar)

        for x in range(1, rows):
            # keep track of the days the product was out of stock

            price_str = str(ws['W%s' % x].value)
            if len(price_str) == 9:
                price_str = price_str[5:]

            if ws['W%s' % x].value is not None and (0 < len(price_str) < 6):

                # fix column with missing values
                find_missing(ws, 'X%s', x)
                find_missing(ws, 'Y%s', x)
                find_missing(ws, 'Z%s', x)

                tos_count = get_tos(ws, (x + 1))
                if tos_count > 0:
                    self.TOS.append(f'{x + 1} was TOS {tos_count}/14 days')

                days_measured = 14 - tos_count
                name = str(ws['U%s' % x].value).strip()
                sku = ws['V%s' % x].value
                price = float(price_str)
                pack_size = self.pack_sizes.get(sku, 1)
                section_name = self.sect_names.get(sku, 'orphan')

                # check for missing prices
                if price == 0:
                    price = self.price_list.get(sku, 0)

                spoiled = get_rate(int(ws['X%s' % x].value), days_measured, x, self.output_box)
                ordered = get_rate(int(ws['Y%s' % x].value), days_measured, x, self.output_box)
                sold = get_rate(int(ws['Z%s' % x].value), days_measured, x, self.output_box)

                dollars = round((price * sold), 2)
                spoil_sales = get_rate(spoiled, sold, x, self.output_box)
                order_sales = get_rate(ordered, sold, x, self.output_box)
                self.data[x].append(section_name)
                self.data[x].append(name)
                self.data[x].append(int(sku))
                self.data[x].append(pack_size)
                self.data[x].append(price)
                self.data[x].append(spoiled)
                self.data[x].append(ordered)
                self.data[x].append(math.ceil(sold))
                self.data[x].append(dollars)
                self.data[x].append(spoil_sales)
                self.data[x].append(order_sales)

        output_msg("Saving dictionary to JSON file.", 50, self.output_box, self.progress_bar)
        with open("source_data/json/sales_data.json", "w") as outfile:
            json.dump(self.data, outfile)

        output_msg("TOS items.", 55, self.output_box, self.progress_bar)

    def build_report(self):
        output_msg("Building Final Report.", 60, self.output_box, self.progress_bar)
        sheet_title = f'454_SECTION_DATA'
        file_name = f'454_{self.report_start_date}_to_{self.report_end_date}'
        file = f'reports/{file_name}.xlsx'

        output_msg("Creating DataFrame.", 60, self.output_box, self.progress_bar)
        df = pd.DataFrame(
            columns=['Section', 'Product', 'SKU', 'Size', 'Price', 'Sold', '$', 'Spoil%', 'Order%'])

        output_msg("Populating DataFrame.", 62, self.output_box, self.progress_bar)
        for key, val in self.data.items():
            df.loc[key] = val[:5] + val[7:]

        output_msg("Writing DataFrame to XLSX.", 70, self.output_box, self.progress_bar)
        writer = pd.ExcelWriter(file, engine='xlsxwriter')
        df.to_excel(writer, sheet_name=sheet_title, index=False)

        output_msg("Styling XLSX.", 71, self.output_box, self.progress_bar)
        wb = writer.book
        money = wb.add_format({'num_format': '$#,##0.00'})
        percent = wb.add_format({'num_format': '0.00%'})

        output_msg("Adjusting Columns.", 78, self.output_box, self.progress_bar)
        ws = writer.sheets[sheet_title]
        ws.set_column('A:A', 10)
        ws.set_column('B:B', 30)
        ws.set_column('C:C', 6)
        ws.set_column('D:D', 5)
        ws.set_column('E:E', 6, money)
        ws.set_column('F:F', 7)
        ws.set_column('G:G', 9, money)
        ws.set_column('H:H', 9, percent)
        ws.set_column('I:I', 10, percent)

        output_msg("Adding AutoFilter.", 88, self.output_box, self.progress_bar)
        ws.autofilter('A1:I1')
        ws.print_title_rows = '1:1'

        output_msg("Setting Sheet to portrait.", 92, self.output_box, self.progress_bar)
        ws.set_portrait()
        ws.repeat_rows(0)
        ws.fit_to_pages(1, 0)
        output_msg("Process Complete.", 100, self.output_box, self.progress_bar)
        writer.save()

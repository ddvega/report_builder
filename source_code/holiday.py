import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side
from source_code.functions import *


class Holiday:
    def __init__(self, prices=None):
        self.prices = prices

    def clean_data(self):
        print('PROCESS INITIATED')
        file = "source_data/xlsx/hsog.xlsx"
        file2 = "source_data/xlsx/holiday_formatted.xlsx"
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = ws.max_row
        end = 10  # number of steps in the process
        print("step 1 of {} complete".format(end))

        copy_col_same_row(rows, ws, 'A%s', 'U%s')  # copy contents of column A to column U

        print("step 2 of {} complete".format(end))

        # fill the SKU column. Fill empty cells with 0
        for cell in range(1, rows):
            y = cell + 1
            a = 'B%s' % y
            b = ws[a].value
            c = str(b)
            d = 'T%s' % cell
            if b is None:
                ws[d] = '0'
            else:
                ws[d] = c[:6]  # trim the data to 6 characters

        print("step 3 of {} complete".format(end))

        copy_col_one_up(rows, ws, 'U%s', 'W%s')
        copy_col_two_up(rows, ws, 'D%s', 'X%s')
        copy_col_two_up(rows, ws, 'E%s', 'Y%s')
        copy_col_two_up(rows, ws, 'F%s', 'Z%s')
        copy_col_two_up(rows, ws, 'G%s', 'AA%s')
        copy_col_two_up(rows, ws, 'I%s', 'AB%s')
        copy_col_two_up(rows, ws, 'J%s', 'AC%s')
        copy_col_two_up(rows, ws, 'K%s', 'AD%s')
        copy_col_two_up(rows, ws, 'L%s', 'AE%s')
        copy_col_two_up(rows, ws, 'M%s', 'AF%s')
        copy_col_two_up(rows, ws, 'N%s', 'AG%s')
        copy_col_two_up(rows, ws, 'O%s', 'AH%s')
        copy_col_two_up(rows, ws, 'P%s', 'AI%s')
        copy_col_two_up(rows, ws, 'Q%s', 'AJ%s')

        print("step 4 of {} complete".format(end))

        # fill empty cells in the same column
        for cell in range(1, rows):
            y = cell + 1
            a = 'X%s' % cell
            b = 'X%s' % y
            c = ws[a].value
            d = ws[b].value
            if d is None or d == '':
                ws[b] = c

        print("step 5 of {} complete".format(end))

        # remove empty space in every cell in column T
        for cell in range(1, rows):
            a = 'T%s' % cell
            b = ws[a].value
            c = str(b)
            if len(c) == 6:
                ws[a] = b[1:]

        print("step 6 of {} complete".format(end))

        copy_col_same_row(rows, ws, 'T%s', 'V%s')  # copy column T to column V

        print("step 7 of {} complete".format(end))

        # remove first 5 spaces in every cell in column W with a length of 9
        for cell in range(1, rows):
            a = 'W%s' % cell
            b = ws[a].value
            c = str(b)
            if len(c) == 9:
                ws[a] = c[5:]

        print("step 8 of {} complete".format(end))

        # convert the cell contents of column W to floats
        for cell in range(1, rows):
            a = 'W%s' % cell
            b = ws[a].value
            c = str(b)
            if b is None:
                continue
            elif len(c) in [1, 2, 3, 4, 5]:
                ws[a] = float(b)

        print("step 9 of {} complete".format(end))

        wb.save(file2)
        df0 = pd.read_excel(file2)

        # select columns to keep and columns to drop
        header = []
        to_drop = []
        for cell in range(0, 36):
            header.append(cell)
            if cell < 20:
                to_drop.append(cell)
        df0.columns = header
        df0 = df0.drop(to_drop, axis='columns')

        # rename column headers
        header2 = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm',
                   'n', 'o', 'p']
        df0.columns = header2

        # filter out all rows in column B that are not greater than 4
        df0 = df0[df0['b'].str.len().gt(4)]

        print("step 10 of {} complete".format(end))

        writer = pd.ExcelWriter(file2, engine='xlsxwriter')
        df0.to_excel(writer, sheet_name="sheet")
        writer.save()

        print('PROCESS COMPLETE')

    def create_report(self):
        increase = decInput()  # projected sales increase for current year
        percentInc = str(round(increase * 100)) + '%'  # show % in footer

        print('PROCESS INITIATED')

        file = "source_data/xlsx/hol_hsog.xlsx"
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        end = 9  # number of steps to complete

        # names of all the sections in the store
        sections = ['drypro', 'grocery', 'frozen', 'wetpro', 'dairy', 'snacks',
                    'cookie.candy', 'deli', 'cheese', 'fresh', 'cereal', 'meat',
                    'wine', 'dfn', 'flowers', 'bakery']

        print("step 1 of {} complete".format(end))

        # get report start date and fill cells on first row
        startDate = get_report_date(ws, 'A%s')
        d0 = startDate
        d1 = d0 + datetime.timedelta(days=1)
        d2 = d0 + datetime.timedelta(days=2)
        d3 = d0 + datetime.timedelta(days=3)
        d4 = d0 + datetime.timedelta(days=4)
        d5 = d0 + datetime.timedelta(days=5)
        d6 = d0 + datetime.timedelta(days=6)
        d7 = d0 + datetime.timedelta(days=7)
        d8 = d0 + datetime.timedelta(days=8)
        d9 = d0 + datetime.timedelta(days=9)
        d10 = d0 + datetime.timedelta(days=10)
        d11 = d0 + datetime.timedelta(days=11)
        d12 = d0 + datetime.timedelta(days=12)

        print("step 2 of {} complete".format(end))

        df0 = pd.read_excel("source_data/xlsx/holiday_formatted.xlsx", header=None)

        df1 = pd.read_excel("source_data/xlsx/sectNames.xlsx", header=None)
        df0.columns = ["product", "sku", "price", d0, d1, d2, d3, d4, d5, d6, d7,
                       d8,
                       d9, d10, d11, d12]  # header for df0

        # search prices dictionary for missing prices in dataframe
        for i in df0.index:
            if df0['price'][i] == 0:
                try:
                    p = prices[str(df0['sku'][i])]
                    if p == 0.01:
                        df0['price'][i] = 0
                    else:
                        df0['price'][i] = p
                except Exception:
                    continue

        df1.columns = ["section", "sku"]  # header for df1
        df0['product'] = df0['product'].str.strip()

        print("step 3 of {} complete".format(end))

        df2 = pd.read_excel("source_data/xlsx/psizes.xlsx", header=None)
        df2.columns = ["sku", "size"]

        print("step 4 of {} complete".format(end))

        df3 = pd.merge(df0, df1, how='left', on='sku')
        df4 = pd.merge(df3, df2, how='left', on='sku')
        df4 = df4[["section", "product", "sku", "size", "price", d0, d1, d2, d3,
                   d4, d5, d6, d7, d8, d9, d10, d11, d12]]
        df4['size'].fillna(1, inplace=True)  # fill NaN cells in column SIZE with 1

        print("step 5 of {} complete".format(end))

        # convert all cells to floats
        dayColumns = [d0, d1, d2, d3, d4, d5, d6, d7, d8, d9, d10, d11, d12]
        for col in df4:
            if col in dayColumns:
                df4[col] = (df4[col].astype(float) +
                            (df4[col].astype(float) * float(increase))) / df4[
                               'size']

        # filename for final report
        finalDraft = "reports/forecast_{}_{}.xlsx".format(d0, d12)

        # reorganize columns on sheet
        df4 = df4[["section", "product", "price", "sku", "size", d0, d1, d2, d3, d4,
                   d5, d6, d7, d8, d9, d10, d11, d12]]
        row_count = len(df4) + 1  # get the number of rows in df4
        writer = pd.ExcelWriter(finalDraft, engine='xlsxwriter')
        sheetName = 'details'

        # print dataframe to excel sheet
        df4.to_excel(writer, sheet_name=sheetName)

        print("step 6 of {} complete".format(end))

        # create a totals sheet with the sum off all day columns
        df5 = df4.groupby(['section'], as_index=False).sum()
        df5.loc[25, "section"] = None
        df5.loc[26, "section"] = 'Totals'
        df5.loc[26, d0] = df5[d0].sum()
        df5.loc[26, d1] = df5[d1].sum()
        df5.loc[26, d2] = df5[d2].sum()
        df5.loc[26, d3] = df5[d3].sum()
        df5.loc[26, d4] = df5[d4].sum()
        df5.loc[26, d5] = df5[d5].sum()
        df5.loc[26, d6] = df5[d6].sum()
        df5.loc[26, d7] = df5[d7].sum()
        df5.loc[26, d8] = df5[d8].sum()
        df5.loc[26, d9] = df5[d9].sum()
        df5.loc[26, d10] = df5[d10].sum()
        df5.loc[26, d11] = df5[d11].sum()
        df5.loc[26, d12] = df5[d12].sum()
        df5 = df5[["section", d0, d1, d2, d3, d4,
                   d5, d6, d7, d8, d9, d10, d11, d12]]
        df5.to_excel(writer, sheet_name='Totals')

        # --------------------------------
        print("step 7 of {} complete".format(end))

        # open workbook to further make styling updates
        wb = writer.book
        ws = writer.sheets[sheetName]
        ws2 = writer.sheets['Totals']
        ws.set_column('B:B', 8)
        ws.set_column('C:C', 20)
        ws.set_column('D:S', 5)
        ws.autofilter('B1:S1')
        ws.set_column('A:A', None, None, {'hidden': True})
        ws.set_landscape()
        green = wb.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
        red = wb.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
        top = wb.add_format({'font_color': '#006100', 'font_size': 3})
        ws.conditional_format('G1:S%s' % row_count, {'type': 'cell',
                                                     'criteria': '=',
                                                     'value': 0, 'format': red})

        # styling for the totals sheet
        ws2.autofilter('B1:O1')
        ws2.set_column('A:A', None, None, {'hidden': True})
        ws2.set_column('B:B', 8)
        ws2.set_column('C:O', 5)

        # save both sheets to workbook
        writer.save()

        print("step 8 of {} complete".format(end))

        # open workbook to further make styling updates to sheet 0 (detailed)
        wrkBook = openpyxl.load_workbook(finalDraft)
        bd = Side(style='thin', color="000000")
        wrkSheet = wrkBook.worksheets[0]
        rows2 = wrkSheet.max_row
        cols = ['A%s', 'B%s', 'C%s', 'D%s', 'E%s', 'F%s', 'G%s', 'H%s', 'I%s',
                'J%s', 'K%s', 'L%s', 'M%s', 'N%s', 'O%s', 'P%s', 'Q%s',
                'R%s', 'S%s']

        # function to create borders around active cells in sheet
        def border_cell1(sheet, rows, column):
            for x in range(1, rows):
                a = column % x
                c = sheet[a]
                c.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        # call function for all cells in all columns
        for cell in cols:
            border_cell1(wrkSheet, rows2, cell)

        # trim date row to show month and day only
        c = ['E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1',
             'P1', 'Q1', 'R1', 'S1']
        for cell in c:
            a = wrkSheet[cell].value
            c = str(a)
            wrkSheet[cell] = c[5:10]

        # set print parameters
        dateS = str(d0)[5:]
        dateE = str(d12)[5:]
        wrkSheet.print_title_rows = '1:1'
        wrkSheet.oddHeader.center.text = '454 forecast from %s to %s' % (
            dateS, dateE)
        wrkSheet.oddFooter.center.text = 'A {} increase has been added to all ' \
                                         ' sales in this time period to reflect ' \
                                         'expected sales growth.'.format(percentInc)

        print("step 9 of {} complete".format(end))

        wrkBook.save(finalDraft)

        print('PROCESS COMPLETE')
